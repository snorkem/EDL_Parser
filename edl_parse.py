import pandas as pd
import pycmx
import logging
import argparse
import yaml
import re
from pathlib import Path
from datetime import datetime
from fnmatch import fnmatch
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from timecode import Timecode

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'./Logs/edl_parser_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)


def parse_edl_to_dataframe(edl_file_path):
    """
    Parse an EDL file using pycmx and convert to pandas DataFrame.

    Args:
        edl_file_path (str): Path to the EDL file

    Returns:
        pandas.DataFrame: DataFrame containing EDL events data
    """
    logger.info(f"Starting EDL parsing for file: {edl_file_path}")

    # Validate file exists
    edl_path = Path(edl_file_path)
    if not edl_path.exists():
        logger.error(f"EDL file not found: {edl_file_path}")
        raise FileNotFoundError(f"EDL file not found: {edl_file_path}")

    logger.info(f"File found, size: {edl_path.stat().st_size} bytes")

    # Pre-parse EDL file to extract motion adapter FPS values and markers
    # pycmx doesn't fully parse M1/M2/M3 lines and markers, so we do a quick raw parse
    motion_fps_map = {}
    marker_map = {}  # Maps event number to list of marker dictionaries
    try:
        with open(edl_file_path, 'r') as f:
            current_event = None
            for line in f:
                line = line.rstrip()
                # Check for event line (starts with digit)
                if line and line[0].isdigit():
                    parts = line.split()
                    if parts:
                        current_event = parts[0]
                # Check for M1/M2/M3 motion adapter line
                elif current_event and line.startswith(('M1 ', 'M2 ', 'M3 ', 'M1\t', 'M2\t', 'M3\t')):
                    parts = line.split()
                    # Find FPS value in the line
                    for part in parts:
                        try:
                            fps = float(part)
                            if 1.0 <= fps <= 240.0:  # Reasonable FPS range
                                motion_fps_map[current_event] = fps
                                break
                        except ValueError:
                            continue
                # Check for marker line (*LOC:)
                elif current_event and line.startswith('*LOC:'):
                    # Parse marker: *LOC: 00:02:23:18 RED     MARKER 4
                    # Remove the *LOC: prefix and split the rest
                    marker_data = line[5:].strip()  # Remove "*LOC:" prefix
                    parts = marker_data.split(None, 2)  # Split into 3 parts: timecode, color, name

                    if len(parts) >= 3:
                        marker_info = {
                            'timecode': parts[0],
                            'color': parts[1],
                            'name': parts[2]
                        }
                    elif len(parts) == 2:
                        marker_info = {
                            'timecode': parts[0],
                            'color': parts[1],
                            'name': ''
                        }
                    else:
                        continue  # Invalid marker format

                    # Add marker to the event's marker list
                    if current_event not in marker_map:
                        marker_map[current_event] = []
                    marker_map[current_event].append(marker_info)

        logger.info(f"Extracted motion FPS data for {len(motion_fps_map)} events")
        logger.info(f"Extracted markers for {len(marker_map)} events")
    except Exception as e:
        logger.warning(f"Could not extract motion FPS data and markers: {str(e)}")
        motion_fps_map = {}
        marker_map = {}

    # Parse EDL file with pycmx
    try:
        with open(edl_file_path, 'r') as f:
            edl = pycmx.parse_cmx3600(f)
        logger.info("Successfully parsed EDL file")
        logger.info(f"EDL Title: {edl.title if edl.title else 'No title'}")
    except (IOError, OSError) as e:
        logger.error(f"Failed to read EDL file: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"Failed to parse EDL file: {str(e)}")
        raise

    # Prepare data for DataFrame
    events_data = []
    events_list = list(edl.events)
    logger.info(f"Total events found: {len(events_list)}")

    # Extract event information
    for idx, event in enumerate(events_list, 1):
        try:
            # Get primary edit (first edit in the event)
            if event.edits:
                edit = event.edits[0]

                # Get motion adapter FPS from pre-parsed map
                # Event numbers in map are zero-padded strings (e.g., '000001'), but pycmx uses integers
                event_key = f"{event.number:06d}"  # Format as 6-digit zero-padded string
                motion_fps = motion_fps_map.get(event_key)
                motion_fps = str(motion_fps) if motion_fps is not None else 'N/A'

                # Get markers from pre-parsed map
                event_markers = marker_map.get(event_key, [])

                # Format marker data - join multiple markers with " | " separator
                if event_markers:
                    marker_timecodes = ' | '.join([m['timecode'] for m in event_markers])
                    marker_colors = ' | '.join([m['color'] for m in event_markers])
                    marker_names = ' | '.join([m['name'] for m in event_markers])
                else:
                    marker_timecodes = ''
                    marker_colors = ''
                    marker_names = ''

                # Format timecodes to ensure 2-digit frames (HH:MM:SS:FF)
                def format_timecode(tc):
                    """Ensure timecode has 2-digit frame count."""
                    if not tc:
                        return 'N/A'
                    tc_str = str(tc)
                    parts = tc_str.split(':')
                    if len(parts) == 4:
                        # Ensure frames are 2 digits
                        parts[3] = parts[3].zfill(2)
                        return ':'.join(parts)
                    return tc_str

                event_data = {
                    'Event #': event.number,
                    'Record In': format_timecode(edit.record_in),
                    'Record Out': format_timecode(edit.record_out),
                    'Clip Name': edit.clip_name if edit.clip_name else 'N/A',
                    'Source File': edit.source_file if edit.source_file else 'N/A',
                    'Reel': edit.source if edit.source else 'N/A',
                    'Source FPS': motion_fps,
                    'Timecode In': format_timecode(edit.source_in),
                    'Timecode Out': format_timecode(edit.source_out),
                    'Transition': edit.transition.name if edit.transition else 'N/A',
                    'Video': 'V' if edit.channels.video else '',
                    'Audio Channels': ', '.join([f'A{ch}' for ch in edit.channels.channels]) if edit.channels.channels else '',
                    'Marker Timecode': marker_timecodes,
                    'Marker Color': marker_colors,
                    'Marker Name': marker_names
                }

                events_data.append(event_data)

                if idx % 10 == 0:
                    logger.info(f"Processed {idx}/{len(events_list)} events")
            else:
                logger.warning(f"Event {event.number} has no edits, skipping")

        except Exception as e:
            logger.error(f"Error processing event {event.number}: {str(e)}")
            continue

    logger.info(f"Successfully processed {len(events_data)} events")

    # Create DataFrame
    df = pd.DataFrame(events_data)

    # Log DataFrame info
    logger.info(f"DataFrame created with shape: {df.shape}")
    logger.info(f"Columns: {', '.join(df.columns)}")

    return df


def load_format_config(config_path):
    """
    Load categorization and formatting configuration from YAML file.

    Args:
        config_path (str): Path to the YAML configuration file

    Returns:
        dict: Configuration dictionary with categories and formatting rules
    """
    logger.info(f"Loading format configuration from: {config_path}")

    try:
        with open(config_path, 'r') as f:
            config = yaml.safe_load(f)

        # Validate basic structure
        if 'categories' not in config:
            raise ValueError("Configuration must contain 'categories' key")

        logger.info(f"Loaded {len(config['categories'])} category rules")
        return config

    except FileNotFoundError:
        logger.error(f"Configuration file not found: {config_path}")
        raise
    except yaml.YAMLError as e:
        logger.error(f"Invalid YAML in configuration file: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"Failed to load configuration: {str(e)}")
        raise


def match_pattern(value, pattern_type, pattern):
    """
    Match a value against a pattern using either glob or regex.

    Args:
        value (str): Value to match against
        pattern_type (str): Either 'glob' or 'regex'
        pattern (str): The pattern to match

    Returns:
        bool: True if pattern matches, False otherwise
    """
    if value is None or value == 'N/A':
        value = ''

    value = str(value)

    try:
        if pattern_type == 'glob':
            return fnmatch(value, pattern)
        elif pattern_type == 'regex':
            return bool(re.search(pattern, value, re.IGNORECASE))
        else:
            logger.warning(f"Unknown pattern type: {pattern_type}")
            return False
    except Exception as e:
        logger.warning(f"Error matching pattern '{pattern}': {str(e)}")
        return False


def categorize_event(event_row, categories):
    """
    Categorize an event based on pattern matching rules.
    Returns ALL matching categories, not just the first one.

    Args:
        event_row (pandas.Series): Row from DataFrame representing an event
        categories (list): List of category configurations

    Returns:
        list: List of tuples [(category_name, category_config), ...] of all matches
              Returns empty list if no matches found
    """
    # Sort categories by priority (lower number = higher priority)
    sorted_categories = sorted(categories, key=lambda c: c.get('priority', 999))

    matched_categories = []

    for category in sorted_categories:
        # Check if any pattern in the category matches (OR logic within a category)
        any_pattern_matches = False

        for pattern_config in category.get('patterns', []):
            field = pattern_config.get('field')
            pattern_type = pattern_config.get('type')
            pattern = pattern_config.get('pattern')

            if not field or not pattern_type or not pattern:
                continue

            # Get field value from event row
            field_value = event_row.get(field, '')

            # Check if pattern matches
            if match_pattern(field_value, pattern_type, pattern):
                any_pattern_matches = True
                break  # Found a match, no need to check other patterns for this category

        if any_pattern_matches:
            matched_categories.append((category['name'], category))

    return matched_categories


def add_categories_to_dataframe(df, config):
    """
    Add category column to DataFrame based on pattern matching rules.
    Supports multiple categories per event, joined with "; ".

    Args:
        df (pandas.DataFrame): DataFrame with EDL events
        config (dict): Configuration dictionary with categories

    Returns:
        pandas.DataFrame: DataFrame with added 'Category' column
    """
    logger.info("Adding categories to events based on pattern matching")

    categories = config.get('categories', [])

    # Add category column
    category_list = []

    for idx, row in df.iterrows():
        matched_categories = categorize_event(row, categories)

        if matched_categories:
            # Join multiple categories with "; "
            category_names = [cat_name for cat_name, _ in matched_categories]
            category_str = "; ".join(category_names)
            category_list.append(category_str)
        else:
            category_list.append('Uncategorized')

    df['Category'] = category_list

    # Log category distribution
    # For multiple categories, count each category separately
    all_categories = []
    for cat_str in category_list:
        if cat_str != 'Uncategorized':
            # Split multi-categories and count each
            all_categories.extend(cat_str.split('; '))
        else:
            all_categories.append(cat_str)

    from collections import Counter
    category_counts = Counter(all_categories)

    logger.info("Category distribution (individual category counts):")
    for cat, count in sorted(category_counts.items(), key=lambda x: -x[1]):
        logger.info(f"  {cat}: {count} events")

    # Also log unique category combinations
    combination_counts = df['Category'].value_counts()
    if len(combination_counts) > 0:
        logger.info("Category combinations:")
        for combo, count in combination_counts.items():
            logger.info(f"  {combo}: {count} events")

    return df


def detect_duplicates(df):
    """
    Detect duplicate events based on timecodes and clip name.

    Args:
        df (pandas.DataFrame): DataFrame with EDL events

    Returns:
        pandas.DataFrame: DataFrame with 'Is_Duplicate' column added
    """
    logger.info("Detecting duplicate events...")

    # Mark duplicates based on Record In, Record Out, and Clip Name
    df['Is_Duplicate'] = df.duplicated(subset=['Record In', 'Record Out', 'Clip Name'], keep='first')

    duplicate_count = df['Is_Duplicate'].sum()
    if duplicate_count > 0:
        logger.warning(f"Found {duplicate_count} duplicate events")
        # Log details of duplicates
        duplicates = df[df['Is_Duplicate']]
        for idx, row in duplicates.iterrows():
            logger.warning(f"  Duplicate: Event #{row['Event #']} - {row['Clip Name']} at {row['Record In']}")
    else:
        logger.info("No duplicate events found")

    return df


def remove_duplicates(df):
    """
    Remove duplicate events from DataFrame.

    Args:
        df (pandas.DataFrame): DataFrame with EDL events

    Returns:
        pandas.DataFrame: DataFrame with duplicates removed
    """
    logger.info("Removing duplicate events...")

    original_count = len(df)
    # Remove duplicates based on Record In, Record Out, and Clip Name
    df = df.drop_duplicates(subset=['Record In', 'Record Out', 'Clip Name'], keep='first')
    removed_count = original_count - len(df)

    if removed_count > 0:
        logger.info(f"Removed {removed_count} duplicate events (kept {len(df)} unique events)")
    else:
        logger.info("No duplicates to remove")

    # Remove the Is_Duplicate column if it exists
    if 'Is_Duplicate' in df.columns:
        df = df.drop('Is_Duplicate', axis=1)

    # Reset index after removing duplicates
    df = df.reset_index(drop=True)

    return df


def generate_statistics(df, config=None):
    """
    Generate comprehensive statistics about the EDL.

    Args:
        df (pandas.DataFrame): DataFrame with EDL events
        config (dict): Optional configuration for category stats

    Returns:
        dict: Dictionary containing various statistics
    """
    logger.info("Generating EDL statistics...")

    from timecode import Timecode

    stats = {}

    # Basic counts
    stats['Total Events'] = len(df)
    stats['Unique Source Files'] = df['Source File'].nunique() if 'Source File' in df.columns else 0
    stats['Unique Clip Names'] = df['Clip Name'].nunique() if 'Clip Name' in df.columns else 0

    # Video vs Audio
    if 'Video' in df.columns:
        video_events = df[df['Video'] == 'V']
        audio_only_events = df[df['Video'] == '']
        stats['Video Events'] = len(video_events)
        stats['Audio Only Events'] = len(audio_only_events)

    # Category distribution
    if 'Category' in df.columns:
        category_counts = df['Category'].value_counts().to_dict()
        stats['Category Distribution'] = category_counts

    # Transition types
    if 'Transition' in df.columns:
        transition_counts = df['Transition'].value_counts().to_dict()
        stats['Transition Types'] = transition_counts

    # FPS distribution
    if 'Source FPS' in df.columns:
        fps_counts = df[df['Source FPS'] != 'N/A']['Source FPS'].value_counts().to_dict()
        if fps_counts:
            stats['FPS Distribution'] = fps_counts

    # Duration analysis (requires timecode calculations)
    try:
        if 'Record In' in df.columns and 'Record Out' in df.columns:
            # Calculate timeline duration (first Record In to last Record Out)
            first_tc = Timecode(30, str(df.iloc[0]['Record In']))
            last_tc = Timecode(30, str(df.iloc[-1]['Record Out']))
            timeline_duration = last_tc.frames - first_tc.frames
            timeline_tc = Timecode(30, frames=timeline_duration)
            stats['Timeline Duration'] = str(timeline_tc)

            # Calculate individual shot durations
            durations = []
            for idx, row in df.iterrows():
                try:
                    rec_in = Timecode(30, str(row['Record In']))
                    rec_out = Timecode(30, str(row['Record Out']))
                    duration_frames = rec_out.frames - rec_in.frames
                    durations.append(duration_frames)
                except:
                    continue

            if durations:
                avg_duration = sum(durations) / len(durations)
                stats['Average Shot Length'] = str(Timecode(30, frames=int(avg_duration)))
                stats['Shortest Shot'] = str(Timecode(30, frames=min(durations)))
                stats['Longest Shot'] = str(Timecode(30, frames=max(durations)))
                stats['Total Shot Frames'] = sum(durations)
    except Exception as e:
        logger.warning(f"Could not calculate duration statistics: {str(e)}")

    # Reel/Tape count
    if 'Reel' in df.columns:
        unique_reels = df[df['Reel'] != 'N/A']['Reel'].nunique()
        stats['Unique Reels'] = unique_reels

    # Duplicates check
    if 'Is_Duplicate' in df.columns:
        stats['Duplicate Events'] = df['Is_Duplicate'].sum()

    logger.info(f"Statistics generated: {len(stats)} metrics")
    return stats


def create_statistics_sheet(stats, workbook_path):
    """
    Add a statistics sheet to the Excel workbook.

    Args:
        stats (dict): Statistics dictionary
        workbook_path (str): Path to Excel file
    """
    logger.info("Adding statistics sheet to Excel...")

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = load_workbook(workbook_path)

        # Create or get Statistics sheet
        if 'Statistics' in wb.sheetnames:
            ws = wb['Statistics']
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet('Statistics', 0)  # Insert as first sheet

        # Add title
        ws['A1'] = 'EDL Statistics Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')

        row = 3

        # Add basic stats
        ws.cell(row, 1, 'Metric').font = Font(bold=True)
        ws.cell(row, 2, 'Value').font = Font(bold=True)
        row += 1

        for key, value in stats.items():
            if key not in ['Category Distribution', 'Transition Types', 'FPS Distribution']:
                ws.cell(row, 1, key)
                ws.cell(row, 2, str(value))
                row += 1

        # Add category distribution if available
        if 'Category Distribution' in stats:
            row += 1
            ws.cell(row, 1, 'Category Distribution').font = Font(bold=True)
            row += 1
            for category, count in stats['Category Distribution'].items():
                ws.cell(row, 1, f"  {category}")
                ws.cell(row, 2, count)
                row += 1

        # Add transition types if available
        if 'Transition Types' in stats:
            row += 1
            ws.cell(row, 1, 'Transition Types').font = Font(bold=True)
            row += 1
            for trans_type, count in stats['Transition Types'].items():
                ws.cell(row, 1, f"  {trans_type}")
                ws.cell(row, 2, count)
                row += 1

        # Add FPS distribution if available
        if 'FPS Distribution' in stats:
            row += 1
            ws.cell(row, 1, 'FPS Distribution').font = Font(bold=True)
            row += 1
            for fps, count in stats['FPS Distribution'].items():
                ws.cell(row, 1, f"  {fps} fps")
                ws.cell(row, 2, count)
                row += 1

        # Auto-size columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        # Apply header fill
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A3'].fill = header_fill
        ws['B3'].fill = header_fill
        ws['A3'].font = Font(color="FFFFFF", bold=True)
        ws['B3'].font = Font(color="FFFFFF", bold=True)

        wb.save(workbook_path)
        logger.info("Statistics sheet added successfully")

    except Exception as e:
        logger.warning(f"Failed to add statistics sheet: {str(e)}")


def apply_excel_formatting(output_path, format_as_table=False, use_colored_rows=False, config=None, df=None, grouped_df=None):
    """
    Apply formatting to the Excel file.

    Args:
        output_path (str): Path to the Excel file
        format_as_table (bool): Format data as Excel table with auto-sized columns
        use_colored_rows (bool): Apply alternating row colors
        config (dict): Optional configuration dictionary for category-based formatting
        df (pandas.DataFrame): Optional DataFrame to map categories to rows
        grouped_df (pandas.DataFrame): Optional grouped events DataFrame for formatting
    """
    if not format_as_table and not use_colored_rows and not config:
        return  # No formatting requested

    logger.info("Applying formatting to Excel file...")

    try:
        # Load the workbook
        wb = load_workbook(output_path)

        # Apply formatting to all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"Formatting sheet: {sheet_name}")

            # Get dimensions
            max_row = ws.max_row
            max_col = ws.max_column

            # Find the Category column index if it exists
            category_col_idx = None
            if config and df is not None:
                for col_idx in range(1, max_col + 1):
                    header_cell = ws.cell(row=1, column=col_idx)
                    if header_cell.value == 'Category':
                        category_col_idx = col_idx
                        break

            # Find the Unique Source Files column for Grouped Events sheet
            source_files_col_idx = None
            if sheet_name == 'Grouped Events' and config and grouped_df is not None:
                for col_idx in range(1, max_col + 1):
                    header_cell = ws.cell(row=1, column=col_idx)
                    if header_cell.value == 'Unique Source Files':
                        source_files_col_idx = col_idx
                        break

            # Auto-size columns if table format requested
            if format_as_table:
                column_widths = {}  # Store column widths for later wrapping check

                for col_idx in range(1, max_col + 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = 0

                    # Check all cells in the column (sample first 100 rows for performance)
                    sample_size = min(max_row, 100)
                    for row_idx in range(1, sample_size + 1):
                        cell = ws[f"{column_letter}{row_idx}"]
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length

                    # Set column width with adjusted calculation
                    # Excel width units are approximately 1.2x character count
                    # Cap at 100 for readability, minimum of 10
                    if max_length > 0:
                        adjusted_width = min(max_length * 1.2 + 2, 100)
                        adjusted_width = max(adjusted_width, 10)
                        ws.column_dimensions[column_letter].width = adjusted_width
                        column_widths[col_idx] = adjusted_width

                # Apply text wrapping to rows where content is clipped
                for row_idx in range(1, max_row + 1):
                    row_needs_wrap = False

                    # Check if any cell in this row has content exceeding column width
                    for col_idx in range(1, max_col + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            cell_length = len(str(cell.value))
                            col_width = column_widths.get(col_idx, 10)
                            # Account for Excel width unit conversion (approx 1.2x)
                            if cell_length > (col_width / 1.2):
                                row_needs_wrap = True
                                break

                    # Apply wrap text to all cells in the row if needed
                    if row_needs_wrap:
                        for col_idx in range(1, max_col + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Find the Is_Duplicate column index if it exists
            duplicate_col_idx = None
            if df is not None and 'Is_Duplicate' in df.columns:
                for col_idx in range(1, max_col + 1):
                    header_cell = ws.cell(row=1, column=col_idx)
                    if header_cell.value == 'Is_Duplicate':
                        duplicate_col_idx = col_idx
                        break

            # Apply duplicate highlighting if Is_Duplicate column exists
            if duplicate_col_idx and df is not None and 'Is_Duplicate' in df.columns:
                logger.info("Applying duplicate highlighting")
                duplicate_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")

                for row_idx in range(2, max_row + 1):
                    is_dup_cell = ws.cell(row=row_idx, column=duplicate_col_idx)
                    if is_dup_cell.value == True or str(is_dup_cell.value).upper() == 'TRUE':
                        # Highlight entire row in red
                        for col_idx in range(1, max_col + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = duplicate_fill

            # Apply rich text formatting to Grouped Events sheet
            elif sheet_name == 'Grouped Events' and config and source_files_col_idx and grouped_df is not None:
                logger.info("Applying category-based rich text formatting to Grouped Events sheet")
                from openpyxl.cell.text import InlineFont
                from openpyxl.cell.rich_text import TextBlock, CellRichText

                # Build category to formatting map
                category_format_map = {}
                for category_config in config.get('categories', []):
                    category_name = category_config['name']
                    formatting = category_config.get('formatting', {})
                    category_format_map[category_name] = formatting

                # Apply formatting row by row
                formatted_count = 0
                for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
                    # Get the grouped event data for this row
                    df_row_idx = row_idx - 2  # DataFrame is 0-indexed
                    if df_row_idx >= len(grouped_df):
                        continue

                    grouped_event_row = grouped_df.iloc[df_row_idx]

                    # Get source file categories metadata
                    source_file_categories = grouped_event_row.get('_source_file_categories', {})
                    if not source_file_categories:
                        continue

                    # Get the cell containing source files
                    source_files_cell = ws.cell(row=row_idx, column=source_files_col_idx)
                    source_files_text = str(source_files_cell.value) if source_files_cell.value else ''

                    if not source_files_text or source_files_text == 'N/A':
                        continue

                    # Split source files by newline
                    source_file_lines = [f.strip() for f in source_files_text.split('\n') if f.strip()]

                    if len(source_file_lines) <= 1:
                        continue  # Skip single-file cells

                    # Build rich text with formatting for each line
                    # KEY: Include newline IN the formatted TextBlock (solution from SOLUTION.md)
                    rich_text_parts = []
                    for i, source_file in enumerate(source_file_lines):
                        # Find the category for this source file
                        category = source_file_categories.get(source_file, '')

                        # Use the first category for formatting (if multiple categories)
                        if category and '; ' in category:
                            primary_category = category.split('; ')[0]
                        else:
                            primary_category = category

                        # Get formatting for this category
                        formatting = category_format_map.get(primary_category, {}) if primary_category else {}

                        # Create inline font with category formatting
                        font_kwargs = {}
                        if 'text_color' in formatting:
                            font_kwargs['color'] = formatting['text_color']
                        if formatting.get('bold', False):
                            font_kwargs['b'] = True
                        if formatting.get('italic', False):
                            font_kwargs['i'] = True

                        # Add newline to the text (except for last file) - SOLUTION FROM SOLUTION.MD
                        if i < len(source_file_lines) - 1:
                            text = source_file + '\n'
                        else:
                            text = source_file

                        # Create TextBlock with formatting
                        if font_kwargs:
                            inline_font = InlineFont(**font_kwargs)
                            text_block = TextBlock(inline_font, text)
                        else:
                            text_block = text

                        rich_text_parts.append(text_block)

                    # Set the cell value to rich text
                    if rich_text_parts:
                        try:
                            source_files_cell.value = CellRichText(*rich_text_parts)
                            formatted_count += 1
                        except Exception as e:
                            logger.warning(f"Could not apply rich text formatting to row {row_idx}: {str(e)}")
                            # Keep original text as fallback
                            pass

                logger.info(f"Applied rich text formatting to {formatted_count} cells in Grouped Events sheet")

            # Apply category-based formatting if config provided
            elif config and category_col_idx and df is not None and 'Category' in df.columns:
                logger.info("Applying category-based formatting")

                # Build category to formatting map
                category_format_map = {}
                for category_config in config.get('categories', []):
                    category_name = category_config['name']
                    formatting = category_config.get('formatting', {})
                    category_format_map[category_name] = formatting

                # Apply formatting row by row
                for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
                    # Get category for this row
                    category_cell = ws.cell(row=row_idx, column=category_col_idx)
                    category = category_cell.value

                    # Handle multiple categories (separated by "; ")
                    # Use the first category for formatting (highest priority)
                    if category and '; ' in str(category):
                        primary_category = str(category).split('; ')[0]
                    else:
                        primary_category = category

                    if primary_category in category_format_map:
                        formatting = category_format_map[primary_category]

                        # Determine which cells to format
                        if formatting.get('entire_row', False):
                            cells_to_format = [ws.cell(row=row_idx, column=c) for c in range(1, max_col + 1)]
                        else:
                            # Only format the category cell
                            cells_to_format = [category_cell]

                        # Apply formatting to each cell
                        for cell in cells_to_format:
                            # Apply fill color
                            if 'cell_color' in formatting:
                                color = formatting['cell_color']
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                            # Apply font formatting
                            font_kwargs = {}
                            if 'text_color' in formatting:
                                font_kwargs['color'] = formatting['text_color']
                            if formatting.get('bold', False):
                                font_kwargs['bold'] = True
                            if formatting.get('italic', False):
                                font_kwargs['italic'] = True

                            if font_kwargs:
                                cell.font = Font(**font_kwargs)

            # Apply alternating row colors if requested (but only if no category formatting)
            elif use_colored_rows:
                # Light gray for alternating rows
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

                # Start from row 2 (skip header)
                for row_idx in range(2, max_row + 1):
                    if row_idx % 2 == 0:  # Even rows
                        for col_idx in range(1, max_col + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill

        # Save the workbook
        wb.save(output_path)
        logger.info("Formatting applied successfully")

    except Exception as e:
        logger.warning(f"Failed to apply formatting: {str(e)}")
        # Don't raise - formatting is optional, file is still valid


def sort_events(df, sort_by='timecode', fps=30):
    """
    Sort events by specified criteria.

    Args:
        df (pandas.DataFrame): DataFrame with EDL events
        sort_by (str): Sort criterion ('timecode', 'clip_name', 'source_file', 'duration', 'category')
        fps (float): Frames per second for duration calculations

    Returns:
        pandas.DataFrame: Sorted DataFrame
    """
    logger.info(f"Sorting events by: {sort_by}")

    df_sorted = df.copy()

    if sort_by == 'timecode':
        # Already sorted by default (Record In)
        logger.info("Events are already in timecode order")
        return df_sorted

    elif sort_by == 'clip_name':
        df_sorted = df_sorted.sort_values('Clip Name', na_position='last')

    elif sort_by == 'source_file':
        df_sorted = df_sorted.sort_values('Source File', na_position='last')

    elif sort_by == 'category':
        if 'Category' in df_sorted.columns:
            df_sorted = df_sorted.sort_values('Category', na_position='last')
        else:
            logger.warning("No Category column found, cannot sort by category")
            return df

    elif sort_by == 'duration':
        # Calculate duration for each event
        try:
            from timecode import Timecode
            durations = []
            for idx, row in df_sorted.iterrows():
                try:
                    rec_in = Timecode(fps, str(row['Record In']))
                    rec_out = Timecode(fps, str(row['Record Out']))
                    duration_frames = rec_out.frames - rec_in.frames
                    durations.append(duration_frames)
                except:
                    durations.append(0)

            df_sorted['_temp_duration'] = durations
            df_sorted = df_sorted.sort_values('_temp_duration', ascending=False)
            df_sorted = df_sorted.drop('_temp_duration', axis=1)
        except Exception as e:
            logger.error(f"Could not sort by duration: {str(e)}")
            return df

    else:
        logger.warning(f"Unknown sort criterion: {sort_by}")
        return df

    # Reset index
    df_sorted = df_sorted.reset_index(drop=True)
    logger.info(f"Sorted {len(df_sorted)} events")

    return df_sorted


def validate_timecode_order(df):
    """
    Validate that events are in sequential timecode order.

    Args:
        df (pandas.DataFrame): DataFrame with EDL events

    Returns:
        tuple: (is_valid, issues_list)
    """
    logger.info("Validating timecode order...")

    issues = []
    from timecode import Timecode

    try:
        for i in range(len(df) - 1):
            current_out = Timecode(30, str(df.iloc[i]['Record Out']))
            next_in = Timecode(30, str(df.iloc[i + 1]['Record In']))

            if next_in.frames < current_out.frames:
                issue = f"Overlap: Event #{df.iloc[i]['Event #']} ends at {current_out} but Event #{df.iloc[i+1]['Event #']} starts at {next_in}"
                issues.append(issue)
                logger.warning(issue)

        if not issues:
            logger.info("Timecode validation passed - all events are in order")
            return True, []
        else:
            logger.warning(f"Timecode validation found {len(issues)} issues")
            return False, issues

    except Exception as e:
        logger.error(f"Timecode validation error: {str(e)}")
        return False, [str(e)]


def group_events_by_time(df, interval_seconds, fps):
    """
    Group events into continuous, sequential time-based intervals.

    Groups are continuous - if Group 1 ends at 00:05:23:08, Group 2 starts at 00:05:23:09.
    Events can appear in multiple groups if they overlap group boundaries.
    Groups are always sequential - no event's Record In is before a previous event's Record Out.

    Args:
        df (pandas.DataFrame): DataFrame with EDL events
        interval_seconds (float): Duration of each time interval in seconds
        fps (int): Frames per second for timecode calculations

    Returns:
        pandas.DataFrame: Grouped events DataFrame
    """
    logger.info(f"Grouping events into continuous {interval_seconds}-second intervals (using Record timecodes at {fps} fps)")

    if df.empty:
        return pd.DataFrame()

    # Determine the starting timecode (first event's Record In)
    first_record_in = str(df.iloc[0]['Record In'])
    start_tc = Timecode(fps, first_record_in)

    # Determine the ending timecode (last event's Record Out)
    last_record_out = str(df.iloc[-1]['Record Out'])
    end_tc = Timecode(fps, last_record_out)

    logger.info(f"Timeline: {first_record_in} to {last_record_out}")

    # Calculate interval in frames
    interval_frames = int(interval_seconds * fps)

    # Create continuous sequential groups
    grouped_events = []
    current_group_start_frames = start_tc.frames
    group_number = 1

    while current_group_start_frames < end_tc.frames:
        # Calculate group end (start + interval)
        current_group_end_frames = current_group_start_frames + interval_frames

        # Don't go beyond the last event's end
        if current_group_end_frames > end_tc.frames:
            current_group_end_frames = end_tc.frames

        # Convert frame numbers to timecodes
        group_start_tc = Timecode(fps, frames=current_group_start_frames)
        group_end_tc = Timecode(fps, frames=current_group_end_frames)

        # Find all events that overlap with this group's time range
        # An event overlaps if: event_start < group_end AND event_end > group_start
        events_in_group = []

        for idx, row in df.iterrows():
            try:
                event_in_str = str(row['Record In'])
                event_out_str = str(row['Record Out'])
                event_in_tc = Timecode(fps, event_in_str)
                event_out_tc = Timecode(fps, event_out_str)

                # Check for overlap
                if event_in_tc.frames < current_group_end_frames and event_out_tc.frames > current_group_start_frames:
                    events_in_group.append(row)

            except Exception as e:
                logger.warning(f"Error processing event {row['Event #']}: {str(e)}")
                continue

        # Create grouped event if there are events in this group
        if events_in_group:
            grouped_event = create_grouped_event(events_in_group, fps, group_number,
                                                 str(group_start_tc), str(group_end_tc))
            grouped_events.append(grouped_event)
            logger.debug(f"Group {group_number}: {group_start_tc} to {group_end_tc} ({len(events_in_group)} events)")

        # Move to next group (start at current end + 1 frame)
        current_group_start_frames = current_group_end_frames + 1
        group_number += 1

    logger.info(f"Created {len(grouped_events)} continuous time-based groups from {len(df)} individual events")

    return pd.DataFrame(grouped_events)


def create_grouped_event(events, fps, group_number=None, group_start=None, group_end=None):
    """
    Create a single grouped event from a list of events.

    Args:
        events (list): List of event rows (pandas Series)
        fps (int): Frames per second for duration calculation
        group_number (int): Optional group number
        group_start (str): Optional group start timecode
        group_end (str): Optional group end timecode

    Returns:
        dict: Grouped event data
    """
    first_event = events[0]
    last_event = events[-1]

    # Collect all clip names
    clip_names = [str(e['Clip Name']) for e in events if e['Clip Name'] != 'N/A']
    unique_clips = len(set(clip_names))

    # Collect all unique source files with their categories
    source_files = []
    source_file_categories = {}  # Map source file to its categories (for formatting)
    for e in events:
        if e['Source File'] != 'N/A':
            source_file = str(e['Source File'])
            source_files.append(source_file)
            # Get category for this event if it exists
            if 'Category' in e and e['Category']:
                if source_file not in source_file_categories:
                    source_file_categories[source_file] = e['Category']

    unique_source_files = sorted(set(source_files))  # Sort for consistent ordering
    source_files_str = '\n'.join(unique_source_files) if unique_source_files else 'N/A'

    # Collect all unique subtitles (if present)
    subtitles = []
    if 'Subtitles' in events[0]:
        for e in events:
            subtitle_text = e.get('Subtitles', '')
            if subtitle_text and isinstance(subtitle_text, str) and subtitle_text.strip():
                # Split by " | " separator (used for multiple subtitles in single event)
                subtitle_parts = [s.strip() for s in subtitle_text.split(' | ') if s.strip()]
                subtitles.extend(subtitle_parts)

        # Remove duplicates while preserving order
        seen = set()
        unique_subtitles = []
        for subtitle in subtitles:
            if subtitle not in seen:
                seen.add(subtitle)
                unique_subtitles.append(subtitle)

        subtitles_str = ' | '.join(unique_subtitles) if unique_subtitles else ''
    else:
        subtitles_str = ''

    # Use group start/end if provided, otherwise use first/last event timecodes
    if group_start and group_end:
        record_in = group_start
        record_out = group_end
    else:
        record_in = first_event['Record In']
        record_out = last_event['Record Out']

    # Calculate duration using timecode
    try:
        record_in_tc = Timecode(fps, str(record_in))
        record_out_tc = Timecode(fps, str(record_out))
        duration_frames = record_out_tc.frames - record_in_tc.frames
        duration_tc = Timecode(fps, frames=duration_frames)
        duration_str = str(duration_tc)
    except:
        duration_str = 'N/A'

    grouped_event = {
        'Group #': group_number if group_number else 1,
        'Event Count': len(events),
        'First Event #': first_event['Event #'],
        'Last Event #': last_event['Event #'],
        'Record In': record_in,
        'Record Out': record_out,
        'Unique Clips': unique_clips,
        'Unique Source Files': source_files_str,
        'Timecode In': first_event['Timecode In'],
        'Timecode Out': last_event['Timecode Out'],
        'Duration': duration_str,
        'Subtitles': subtitles_str,
        '_source_file_categories': source_file_categories  # For formatting (not exported to Excel)
    }

    return grouped_event


def export_to_excel(df, output_path, format_as_table=False, use_colored_rows=False, grouped_df=None, config=None):
    """
    Export DataFrame to Excel file.

    Args:
        df (pandas.DataFrame): DataFrame to export
        output_path (str): Output Excel file path
        format_as_table (bool): Format data as Excel table with auto-sized columns
        use_colored_rows (bool): Apply alternating row colors
        grouped_df (pandas.DataFrame): Optional grouped events DataFrame for second sheet
        config (dict): Optional configuration dictionary for category-based formatting
    """
    logger.info(f"Exporting DataFrame to Excel: {output_path}")

    try:
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write main sheet
            df.to_excel(writer, sheet_name='Events', index=False)

            # Write grouped sheet if provided
            if grouped_df is not None and not grouped_df.empty:
                # Exclude the _source_file_categories column from Excel export
                export_cols = [col for col in grouped_df.columns if not col.startswith('_')]
                grouped_df[export_cols].to_excel(writer, sheet_name='Grouped Events', index=False)
                logger.info(f"Added 'Grouped Events' sheet with {len(grouped_df)} groups")

        logger.info(f"Successfully exported to {output_path}")

        # Apply formatting if requested (pass grouped_df for category formatting)
        apply_excel_formatting(output_path, format_as_table, use_colored_rows, config, df, grouped_df)

        logger.info(f"File size: {Path(output_path).stat().st_size} bytes")
    except (IOError, OSError, PermissionError) as e:
        logger.error(f"Failed to write Excel file: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"Failed to export to Excel: {str(e)}")
        raise


def main():
    """Main execution function."""
    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description='Convert EDL files to Excel spreadsheets with comprehensive logging.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python edl_parse.py input.edl output.xlsx
  python edl_parse.py input.edl output.xlsx --config format_config.yaml
  python edl_parse.py input.edl output.xlsx --no-table --no-colored
  python edl_parse.py input.edl output.xlsx --group 15 --fps 24 --config format_config.yaml

Note: Table formatting and colored rows are enabled by default.
      Use --no-table and --no-colored to disable them.
        """
    )

    parser.add_argument(
        'input_edl',
        nargs='?',
        help='Path to the input EDL file (not required when using --merge or --compare)'
    )

    parser.add_argument(
        'output_xlsx',
        nargs='?',
        help='Path to the output Excel file (optional, defaults to input filename with .xlsx extension)'
    )

    parser.add_argument(
        '--no-table',
        action='store_false',
        dest='table',
        help='Disable table formatting with auto-sized columns (enabled by default)'
    )

    parser.add_argument(
        '--no-colored',
        action='store_false',
        dest='colored',
        help='Disable alternating row colors (enabled by default when not using --config)'
    )

    parser.add_argument(
        '--group',
        type=float,
        metavar='SECONDS',
        help='Group events with gaps smaller than SECONDS into a second sheet (e.g., --group 15). Requires --fps.'
    )

    parser.add_argument(
        '--fps',
        type=float,
        metavar='FPS',
        help='Frames per second for timecode calculations (required if using --group). Common values: 23.976, 24, 25, 29.97, 30, 50, 59.94, 60'
    )

    parser.add_argument(
        '--config',
        type=str,
        metavar='CONFIG_FILE',
        help='Path to JSON configuration file for categorization and formatting rules'
    )

    parser.add_argument(
        '--remove-duplicates',
        action='store_true',
        help='Remove duplicate events (same timecode and clip name)'
    )

    parser.add_argument(
        '--highlight-duplicates',
        action='store_true',
        help='Highlight duplicate events in red (does not remove them)'
    )

    parser.add_argument(
        '--stats',
        action='store_true',
        help='Add statistics sheet to Excel output'
    )

    parser.add_argument(
        '--stats-only',
        action='store_true',
        help='Export only statistics (no events data)'
    )

    parser.add_argument(
        '--sort-by',
        type=str,
        choices=['timecode', 'clip_name', 'source_file', 'duration', 'category'],
        help='Sort events by specified criterion'
    )

    parser.add_argument(
        '--validate',
        action='store_true',
        help='Validate timecode order and report issues'
    )

    parser.add_argument(
        '--filter',
        type=str,
        metavar='EXPRESSION',
        help='Filter events using pandas query expression (e.g., "Category == \'A-Camera\'")'
    )

    parser.add_argument(
        '--search',
        type=str,
        metavar='TERM',
        help='Search for events matching term (glob pattern)'
    )

    parser.add_argument(
        '--search-field',
        type=str,
        help='Specific field to search in (default: all text fields)'
    )

    parser.add_argument(
        '--search-regex',
        action='store_true',
        help='Use regex for search instead of glob'
    )

    parser.add_argument(
        '--merge',
        type=str,
        nargs='+',
        metavar='EDL_FILE',
        help='Merge multiple EDL files'
    )

    parser.add_argument(
        '--output',
        '-o',
        type=str,
        metavar='OUTPUT_FILE',
        help='Output Excel file path (alternative to positional output_xlsx argument)'
    )

    parser.add_argument(
        '--subtitle-file',
        '--stl-file',  # Keep old name for backwards compatibility
        type=str,
        metavar='SUBTITLE_FILE',
        dest='subtitle_file',
        help='Path to subtitle file (.stl or .srt) to merge with EDL events (requires --merge)'
    )

    parser.add_argument(
        '--subtitle-fps',
        '--stl-fps',  # Keep old name for backwards compatibility
        type=float,
        metavar='FPS',
        dest='subtitle_fps',
        help='Frame rate for subtitle parsing. STL: auto-detects if not specified (user value overrides). SRT: defaults to 30 if not specified. Common values: 23.976, 24, 25, 29.97, 30'
    )

    parser.add_argument(
        '--subtitle-start-time',
        '--stl-start-time',  # Keep old name for backwards compatibility
        type=str,
        metavar='TIMECODE',
        dest='subtitle_start_time',
        help='Record In timecode (HH:MM:SS:FF) where the first subtitle should align in the EDL timeline. The offset is calculated automatically based on the first subtitle timecode.'
    )

    parser.add_argument(
        '--split-by-category',
        action='store_true',
        help='Split events by category into separate Excel files'
    )

    parser.add_argument(
        '--split-output-dir',
        type=str,
        default='./split_output',
        help='Output directory for split files (default: ./split_output)'
    )

    parser.add_argument(
        '--compare',
        type=str,
        nargs=2,
        metavar=('ORIGINAL_EDL', 'REVISED_EDL'),
        help='Compare two EDL files and generate changelog'
    )

    parser.add_argument(
        '--changelog-output',
        type=str,
        help='Output path for changelog report (default: changelog.xlsx)'
    )

    args = parser.parse_args()

    # Validate that either input_edl OR --merge/--compare is provided
    if not args.input_edl and not args.merge and not args.compare:
        parser.error('input_edl is required unless using --merge or --compare')

    # Validate that --fps is provided when --group is used
    if args.group is not None and args.fps is None:
        parser.error('--fps is required when using --group')

    # Set input and output paths
    edl_file_path = args.input_edl

    # Determine output path (check --output flag first, then positional arg, then generate default)
    if args.output:
        output_file_path = args.output
    elif args.output_xlsx:
        output_file_path = args.output_xlsx
    else:
        # Only generate output path from input_edl if it's provided
        if edl_file_path:
            input_path = Path(edl_file_path)
            output_file_path = str(input_path.parent / f"{input_path.stem}_output.xlsx")
        else:
            # Default output path for --merge or --compare modes
            output_file_path = "output.xlsx"

    logger.info("=" * 60)
    logger.info("EDL to Excel Converter - Starting")
    logger.info("=" * 60)

    try:
        # Handle special modes first

        # EDL Comparison mode
        if args.compare:
            from utils.edl_advanced import compare_edls, create_changelog_report
            logger.info(f"Comparing EDLs: {args.compare[0]} vs {args.compare[1]}")
            comparison = compare_edls(args.compare[0], args.compare[1])
            changelog_path = args.changelog_output if args.changelog_output else 'changelog.xlsx'
            create_changelog_report(comparison, changelog_path)
            logger.info(f"Changelog report created: {changelog_path}")
            logger.info("=" * 60)
            logger.info("EDL Comparison - Completed Successfully")
            logger.info("=" * 60)
            return comparison

        # EDL Merge mode
        if args.merge:
            from utils.edl_advanced import merge_edls
            logger.info(f"Merging {len(args.merge)} EDL files (interleave mode)")

            # Pass subtitle parameters if provided
            subtitle_file = args.subtitle_file if hasattr(args, 'subtitle_file') else None
            subtitle_fps = args.subtitle_fps if hasattr(args, 'subtitle_fps') else None
            subtitle_start_time = args.subtitle_start_time if hasattr(args, 'subtitle_start_time') else None

            df = merge_edls(args.merge, subtitle_file=subtitle_file, subtitle_fps=subtitle_fps, subtitle_start_time=subtitle_start_time)
            edl_file_path = "merged_edl"  # For logging purposes
        else:
            # Normal mode: parse single EDL
            logger.info(f"Input EDL: {edl_file_path}")
            df = parse_edl_to_dataframe(edl_file_path)

        logger.info(f"Output Excel: {output_file_path}")
        if args.config:
            logger.info(f"Config file: {args.config}")

        # Load configuration if provided
        config = None
        if args.config:
            config = load_format_config(args.config)

        # Apply categorization if config provided
        if config:
            df = add_categories_to_dataframe(df, config)

        # Handle duplicates
        if args.highlight_duplicates and not args.remove_duplicates:
            df = detect_duplicates(df)
        elif args.remove_duplicates:
            # Detect first to log them, then remove
            df = detect_duplicates(df)
            df = remove_duplicates(df)

        # Sort events if requested
        if args.sort_by:
            sort_fps = args.fps if args.fps else 30
            df = sort_events(df, args.sort_by, sort_fps)

        # Validate timecode order if requested
        if args.validate:
            is_valid, issues = validate_timecode_order(df)
            if not is_valid:
                logger.warning("Timecode validation issues found:")
                for issue in issues:
                    logger.warning(f"  - {issue}")

        # Apply filtering if requested
        if args.filter:
            from utils.edl_advanced import filter_events
            df = filter_events(df, args.filter)

        # Apply search if requested
        if args.search:
            from utils.edl_advanced import search_events
            df = search_events(df, args.search, args.search_field, args.search_regex)

        # Generate statistics if requested
        stats = None
        if args.stats or args.stats_only:
            stats = generate_statistics(df, config)
            # Log statistics
            logger.info("\n=== EDL Statistics ===")
            for key, value in stats.items():
                if key not in ['Category Distribution', 'Transition Types', 'FPS Distribution']:
                    logger.info(f"{key}: {value}")

        # If stats-only, skip event export
        if args.stats_only:
            # Create a minimal Excel with just statistics
            import pandas as pd
            empty_df = pd.DataFrame()
            export_to_excel(empty_df, output_file_path, False, False, None, config)
            if stats:
                create_statistics_sheet(stats, output_file_path)
        else:
            # Display preview
            logger.info("\nDataFrame Preview (first 5 rows):")
            logger.info("\n" + df.head().to_string())

            # Group events if requested
            grouped_df = None
            if args.group is not None:
                grouped_df = group_events_by_time(df, args.group, args.fps)
                if not grouped_df.empty:
                    logger.info("\nGrouped Events Preview (first 5 rows):")
                    logger.info("\n" + grouped_df.head().to_string())

            # Export to Excel
            export_to_excel(df, output_file_path, args.table, args.colored, grouped_df, config)

            # Add statistics sheet if requested
            if args.stats and stats:
                create_statistics_sheet(stats, output_file_path)

        # Split by category if requested
        if args.split_by_category:
            from utils.edl_advanced import split_by_category
            if 'Category' not in df.columns:
                logger.warning("Cannot split by category: no Category column found. Use --config to add categories.")
            else:
                split_files = split_by_category(df, args.split_output_dir, 'Category')
                logger.info(f"Split into {len(split_files)} category files in {args.split_output_dir}")

        logger.info("=" * 60)
        logger.info("EDL to Excel Converter - Completed Successfully")
        logger.info("=" * 60)

        return df

    except Exception as e:
        logger.error(f"Program failed with error: {str(e)}")
        logger.info("=" * 60)
        logger.info("EDL to Excel Converter - Failed")
        logger.info("=" * 60)
        raise


if __name__ == "__main__":
    main()
