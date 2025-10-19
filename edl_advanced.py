"""
Advanced EDL processing functions - Filtering, Merging, Comparison
"""

import pandas as pd
import logging
from pathlib import Path
import re
from fnmatch import fnmatch

logger = logging.getLogger(__name__)


def filter_events(df, filter_expr):
    """
    Filter events based on a filter expression.

    Args:
        df (pandas.DataFrame): DataFrame with EDL events
        filter_expr (str): Filter expression (e.g., "category == 'A-Camera' and duration > 5")

    Returns:
        pandas.DataFrame: Filtered DataFrame
    """
    logger.info(f"Applying filter: {filter_expr}")

    try:
        df_filtered = df.query(filter_expr)
        logger.info(f"Filter result: {len(df_filtered)} events (from {len(df)} total)")
        return df_filtered
    except Exception as e:
        logger.error(f"Filter error: {str(e)}")
        return df


def search_events(df, search_term, field=None, use_regex=False):
    """
    Search for events matching a term.

    Args:
        df (pandas.DataFrame): DataFrame with EDL events
        search_term (str): Search term (glob or regex)
        field (str): Specific field to search (None for all text fields)
        use_regex (bool): Use regex instead of glob

    Returns:
        pandas.DataFrame: DataFrame with matching events
    """
    logger.info(f"Searching for: {search_term}" + (f" in field: {field}" if field else " in all fields"))

    text_fields = ['Clip Name', 'Source File', 'Reel']
    if field:
        text_fields = [field] if field in df.columns else text_fields

    matches = pd.Series([False] * len(df))

    for field_name in text_fields:
        if field_name in df.columns:
            for idx, value in df[field_name].items():
                value_str = str(value) if value else ''
                if use_regex:
                    if re.search(search_term, value_str, re.IGNORECASE):
                        matches[idx] = True
                else:
                    if fnmatch(value_str.lower(), search_term.lower()):
                        matches[idx] = True

    df_result = df[matches].copy()
    logger.info(f"Found {len(df_result)} matching events")
    return df_result


def merge_edls(edl_file_list, subtitle_file=None, subtitle_fps=None, subtitle_start_time=None):
    """
    Merge multiple EDL files into a single DataFrame, with optional subtitle matching.
    Can also be used to add subtitles to a single EDL file.

    EDL files are merged and sorted by Record In timecode (interleave mode).
    Supports both STL (.stl) and SRT (.srt) subtitle files.

    Args:
        edl_file_list (list): List of EDL file paths (can be a single file)
        subtitle_file (str): Optional path to subtitle file (.stl or .srt)
        subtitle_fps (float): Optional FPS for subtitle parsing (STL: auto-detects if None, SRT: default 30)
        subtitle_start_time (str): Optional Record In timecode (HH:MM:SS:FF) where first subtitle should align

    Returns:
        pandas.DataFrame: Merged DataFrame with optional 'Subtitles' column
    """
    from edl_parse import parse_edl_to_dataframe
    from timecode import Timecode

    if len(edl_file_list) == 1:
        subtitle_type = None
        if subtitle_file:
            if subtitle_file.lower().endswith('.stl'):
                subtitle_type = "STL"
            elif subtitle_file.lower().endswith('.srt'):
                subtitle_type = "SRT"
        logger.info(f"Processing single EDL file" + (f" with {subtitle_type} subtitles" if subtitle_type else ""))
    else:
        logger.info(f"Merging {len(edl_file_list)} EDL files (interleave mode)")

    dfs = []
    for edl_file in edl_file_list:
        try:
            df = parse_edl_to_dataframe(edl_file)
            dfs.append(df)
            logger.info(f"  Loaded: {edl_file} ({len(df)} events)")
        except Exception as e:
            logger.error(f"  Failed to load {edl_file}: {str(e)}")

    if not dfs:
        logger.error("No EDL files could be loaded")
        return pd.DataFrame()

    # Merge and sort by Record In timecode (interleave mode)
    merged_df = pd.concat(dfs, ignore_index=True)
    merged_df = merged_df.sort_values('Record In')

    # Renumber events
    merged_df['Event #'] = range(1, len(merged_df) + 1)

    if len(edl_file_list) == 1:
        logger.info(f"Processed {len(merged_df)} events from single EDL")
    else:
        logger.info(f"Merged result: {len(merged_df)} total events")

    # Merge subtitles if provided (STL or SRT)
    if subtitle_file:
        try:
            # Detect subtitle file type
            subtitle_path = Path(subtitle_file)
            file_ext = subtitle_path.suffix.lower()

            if file_ext == '.stl':
                # Parse STL file
                from stl_parser import parse_stl_file, match_subtitles_to_events

                logger.info(f"Parsing STL file: {subtitle_file}")
                subtitles, effective_fps = parse_stl_file(subtitle_file, fps=subtitle_fps)

                logger.info(f"Matching {len(subtitles)} STL subtitles to {len(merged_df)} events at {effective_fps} fps")
                merged_df = match_subtitles_to_events(merged_df, subtitles, effective_fps, stl_start_time=subtitle_start_time)

            elif file_ext == '.srt':
                # Parse SRT file
                from srt_parser import parse_srt_file, match_subtitles_to_events

                # SRT uses user-specified FPS (default 30 if not specified)
                effective_fps = subtitle_fps if subtitle_fps else 30

                logger.info(f"Parsing SRT file: {subtitle_file}")
                subtitles, effective_fps = parse_srt_file(subtitle_file, fps=effective_fps)

                logger.info(f"Matching {len(subtitles)} SRT subtitles to {len(merged_df)} events at {effective_fps} fps")
                merged_df = match_subtitles_to_events(merged_df, subtitles, effective_fps, srt_start_time=subtitle_start_time)

            else:
                logger.error(f"Unsupported subtitle format: {file_ext}. Supported formats: .stl, .srt")
                logger.warning("Continuing without subtitle data")

        except Exception as e:
            logger.error(f"Failed to process subtitle file: {str(e)}")
            logger.warning("Continuing without subtitle data")

    return merged_df


def split_by_category(df, output_dir, category_col='Category', format='xlsx'):
    """
    Split events by category into separate Excel files.

    Args:
        df (pandas.DataFrame): DataFrame with EDL events
        output_dir (str): Output directory
        category_col (str): Column to split by
        format (str): Output format (only 'xlsx' supported)

    Returns:
        dict: Map of category -> output file path
    """
    logger.info(f"Splitting events by {category_col}")

    if category_col not in df.columns:
        logger.error(f"Column '{category_col}' not found in DataFrame")
        return {}

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    output_files = {}

    for category in df[category_col].unique():
        df_category = df[df[category_col] == category].copy()
        category_safe = re.sub(r'[^\w\-_]', '_', category)

        output_path = output_dir / f"{category_safe}.xlsx"
        df_category.to_excel(output_path, index=False)

        output_files[category] = str(output_path)
        logger.info(f"  {category}: {len(df_category)} events -> {output_path}")

    logger.info(f"Split into {len(output_files)} files")
    return output_files


def compare_edls(edl_file1, edl_file2):
    """
    Compare two EDL files and identify changes.

    Args:
        edl_file1 (str): Original EDL file path
        edl_file2 (str): Revised EDL file path

    Returns:
        dict: Dictionary with added, removed, and modified events
    """
    from edl_parse import parse_edl_to_dataframe

    logger.info(f"Comparing EDLs: {edl_file1} vs {edl_file2}")

    df1 = parse_edl_to_dataframe(edl_file1)
    df2 = parse_edl_to_dataframe(edl_file2)

    logger.info(f"  Original: {len(df1)} events")
    logger.info(f"  Revised: {len(df2)} events")

    # Create comparison keys
    df1['_compare_key'] = df1['Record In'] + '_' + df1['Record Out'] + '_' + df1['Clip Name']
    df2['_compare_key'] = df2['Record In'] + '_' + df2['Record Out'] + '_' + df2['Clip Name']

    keys1 = set(df1['_compare_key'])
    keys2 = set(df2['_compare_key'])

    # Find added and removed
    added_keys = keys2 - keys1
    removed_keys = keys1 - keys2
    common_keys = keys1 & keys2

    added_events = df2[df2['_compare_key'].isin(added_keys)].copy()
    removed_events = df1[df1['_compare_key'].isin(removed_keys)].copy()

    # Find modified (same timecode, different source)
    modified_events = []
    for key in common_keys:
        row1 = df1[df1['_compare_key'] == key].iloc[0]
        row2 = df2[df2['_compare_key'] == key].iloc[0]

        if row1['Source File'] != row2['Source File']:
            modified_events.append({
                'Event #': row2['Event #'],
                'Record In': row2['Record In'],
                'Record Out': row2['Record Out'],
                'Old Clip': row1['Clip Name'],
                'New Clip': row2['Clip Name'],
                'Old Source': row1['Source File'],
                'New Source': row2['Source File'],
                'Change Type': 'Replaced'
            })

    # Clean up
    df1 = df1.drop('_compare_key', axis=1)
    df2 = df2.drop('_compare_key', axis=1)
    if len(added_events) > 0:
        added_events = added_events.drop('_compare_key', axis=1)
    if len(removed_events) > 0:
        removed_events = removed_events.drop('_compare_key', axis=1)

    logger.info(f"  Added: {len(added_events)} events")
    logger.info(f"  Removed: {len(removed_events)} events")
    logger.info(f"  Modified: {len(modified_events)} events")

    return {
        'added': added_events,
        'removed': removed_events,
        'modified': pd.DataFrame(modified_events) if modified_events else pd.DataFrame(),
        'original': df1,
        'revised': df2
    }


def create_changelog_report(comparison, output_path):
    """
    Create changelog Excel report from EDL comparison.

    Args:
        comparison (dict): Output from compare_edls()
        output_path (str): Output Excel file path
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font

    logger.info(f"Creating changelog report: {output_path}")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write summary sheet
        summary_data = {
            'Metric': ['Total Events (Original)', 'Total Events (Revised)', 'Added Events', 'Removed Events', 'Modified Events'],
            'Count': [
                len(comparison['original']),
                len(comparison['revised']),
                len(comparison['added']),
                len(comparison['removed']),
                len(comparison['modified'])
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

        # Write detailed sheets
        if len(comparison['added']) > 0:
            comparison['added'].to_excel(writer, sheet_name='Added Events', index=False)

        if len(comparison['removed']) > 0:
            comparison['removed'].to_excel(writer, sheet_name='Removed Events', index=False)

        if len(comparison['modified']) > 0:
            comparison['modified'].to_excel(writer, sheet_name='Modified Events', index=False)

    # Apply color formatting
    try:
        wb = load_workbook(output_path)

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # Color added events green
        if 'Added Events' in wb.sheetnames:
            ws = wb['Added Events']
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).fill = green_fill

        # Color removed events red
        if 'Removed Events' in wb.sheetnames:
            ws = wb['Removed Events']
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).fill = red_fill

        # Color modified events yellow
        if 'Modified Events' in wb.sheetnames:
            ws = wb['Modified Events']
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).fill = yellow_fill

        wb.save(output_path)
        logger.info("Changelog report created with color coding")

    except Exception as e:
        logger.warning(f"Could not apply color formatting: {str(e)}")

    logger.info(f"Changelog report saved: {output_path}")
